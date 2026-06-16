"""Detect 고정값 columns from a Marketo template and extract their values from
a prior filled-in result.

The template marks each column's source on row 2:
    - '고정값'         -> constant for the entire event (e.g. Country, Import Owner)
    - '행사명 영문'    -> event name (Import Name)
    - '행사 날짜'      -> event date (Initial Response Date)
    - '마케팅 제공'    -> marketing-provided constant (e.g. SFDC Campaign ID)
    - '드롭다운'       -> dropdown selection (per-event constant)
    - '[등록정보] ...' -> per-row from source data
    - '공란'           -> always blank

Workflow:
    constant_cols = detect_constant_columns(template_path)
    values = extract_first_row_values(prior_result_path, constant_cols)
    # values: {'Country': 'Korea, Republic of', 'Import Owner': '...', ...}

Use these values as EVENT_CONSTANTS in your converter so they are pulled from a
prior filled template instead of being re-typed each event.
"""
import openpyxl


# Row 2 markers that indicate a column whose value is constant for the event.
# Per-row markers (start with '[등록정보]') and '공란' are excluded.
CONSTANT_MARKERS = {'고정값', '행사명 영문', '행사 날짜', '마케팅 제공', '드롭다운'}


def detect_constant_columns(template_file, sheet='Sheet1',
                            marker_row=2, header_row=3):
    """Return list of (col_index_0based, marketo_column_name, marker_text)
    for every column whose row-2 marker indicates a per-event constant."""
    wb = openpyxl.load_workbook(template_file, data_only=True)
    ws = wb[sheet]
    result = []
    for col in range(1, ws.max_column + 1):
        marker = ws.cell(row=marker_row, column=col).value
        if marker is None:
            continue
        marker_text = str(marker).strip()
        if marker_text in CONSTANT_MARKERS:
            header = ws.cell(row=header_row, column=col).value
            if header:
                result.append((col - 1, str(header).strip(), marker_text))
    return result


def extract_first_row_values(prior_result_file, constant_cols,
                             sheet='Sheet1', data_row=5):
    """Pull row `data_row` values for the constant columns from a prior result.

    Args:
        prior_result_file: path to a previously filled Marketo import .xlsx.
        constant_cols: output of detect_constant_columns().
        sheet: worksheet name.
        data_row: 1-based row index of the first data row (default 5).

    Returns:
        dict[marketo_column_name -> value]. Skips columns whose first-row
        cell is empty so callers can fall back to defaults.
    """
    wb = openpyxl.load_workbook(prior_result_file, data_only=True)
    ws = wb[sheet]
    values = {}
    for col_idx, header, _marker in constant_cols:
        cell_value = ws.cell(row=data_row, column=col_idx + 1).value
        if cell_value is not None and str(cell_value).strip() != '':
            values[header] = cell_value
    return values


def build_event_constants(template_file, prior_result_file):
    """Convenience: return EVENT_CONSTANTS dict ready to splat into a row."""
    cols = detect_constant_columns(template_file)
    return extract_first_row_values(prior_result_file, cols)
