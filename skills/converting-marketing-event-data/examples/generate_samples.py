"""Generate anonymized sample workbooks for the repo.

Run once to seed examples/. The output files are committed to the repo so
new users can run the Quick Start without real PII.

Usage:
    python examples/generate_samples.py
"""
import os
import shutil

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


HERE = os.path.dirname(os.path.abspath(__file__))


SOURCE_ROWS = [
    # 한글 이름 (2자, 3자, 2자 성씨 포함), POC/상담/세미나 follow-up 신호 다양화
    {'신청일': '2026.05.21 09:00', '성명': '김민수',     '이메일': 'minsu.kim@example-corp.kr',  '연락처': '010-1234-5678', '회사명': '예시전자',     '영문회사명': None,           '직책(직급)': '책임',          '부서': '연구소',     '초청경로': '인터넷',  '유입경로': 'metaad1fb'},
    {'신청일': '2026.05.21 10:30', '성명': '이지영',     '이메일': 'jiyoung.lee@samplebank.kr',   '연락처': '010-2345-6789', '회사명': '샘플은행',     '영문회사명': 'Sample Bank',  '직책(직급)': '팀장',          '부서': '디지털혁신팀','초청경로': '메일',    '유입경로': 'POC 신청'},
    {'신청일': '2026.05.22 11:15', '성명': '남궁민지',   '이메일': 'minji.namgung@demohealth.kr', '연락처': '010-3456-7890', '회사명': '데모헬스케어', '영문회사명': None,           '직책(직급)': '상무',          '부서': '경영지원',   '초청경로': '추천',    '유입경로': '미팅 요청'},
    {'신청일': '2026.05.22 14:00', '성명': '박재현',     '이메일': 'jaehyun.park@testlogi.com',   '연락처': 1067891234,      '회사명': '테스트물류',   '영문회사명': None,           '직책(직급)': '대리',          '부서': '운영팀',     '초청경로': 'email',  '유입경로': '자료요청'},
    {'신청일': '2026.05.23 09:45', '성명': 'John Doe',  '이메일': 'john.doe@globalmfg.com',      '연락처': '+82-10-4567-8901','회사명': '글로벌제조',  '영문회사명': 'Global Mfg',   '직책(직급)': 'Senior Engineer', '부서': 'Production', '초청경로': '광고',    '유입경로': '세미나 신청'},
    {'신청일': '2026.05.23 13:20', '성명': '최수진',     '이메일': 'sujin.choi@retailco.kr',      '연락처': '010-5678-9012', '회사명': '리테일컴퍼니', '영문회사명': None,           '직책(직급)': '과장',          '부서': '마케팅',     '초청경로': '뉴스레터','유입경로': '아니오'},
    {'신청일': '2026.05.23 15:00', '성명': '정대표',     '이메일': 'ceo@samplegov.kr',            '연락처': '010-6789-0123', '회사명': '샘플공공기관', '영문회사명': None,           '직책(직급)': '대표',          '부서': '경영진',     '초청경로': '지인 추천','유입경로': '기업방문교육 요청'},
    {'신청일': '2026.05.24 10:00', '성명': 'JaneSmith', '이메일': 'jane@telcoexample.kr',        '연락처': '010-7890-1234', '회사명': '예시텔레콤',   '영문회사명': None,           '직책(직급)': 'Director',      '부서': 'Network',    '초청경로': 'LinkedIn','유입경로': '무응답'},
    {'신청일': '2026.05.24 11:30', '성명': '한지원',     '이메일': '',                            '연락처': '010-8901-2345', '회사명': '잘못된이메일', '영문회사명': None,           '직책(직급)': '사원',          '부서': '인사',       '초청경로': '없음',    '유입경로': None},
    {'신청일': '2026.05.24 16:45', '성명': '윤서준',     '이메일': 'seojun.yoon@sampleins.kr',    '연락처': '010-9012-3456', '회사명': '샘플화재보험', '영문회사명': 'Sample Fire',  '직책(직급)': '부장',          '부서': '영업본부',   '초청경로': '메일링', '유입경로': '상담 요청'},
]

COMPANY_ROWS = [
    {'No.': 1, '한글 회사명': '예시전자',     '영문 회사명': 'Example Electronics', 'Industry': 'Manufacturing'},
    {'No.': 2, '한글 회사명': '샘플은행',     '영문 회사명': 'Sample Bank',         'Industry': 'Banking and Financial Services'},
    {'No.': 3, '한글 회사명': '데모헬스케어', '영문 회사명': 'Demo Healthcare',     'Industry': 'Healthcare'},
    {'No.': 4, '한글 회사명': '글로벌제조',   '영문 회사명': 'Global Mfg',          'Industry': 'Manufacturing'},
    {'No.': 5, '한글 회사명': '리테일컴퍼니', '영문 회사명': 'Retail Company',      'Industry': 'Retail'},
    {'No.': 6, '한글 회사명': '예시텔레콤',   '영문 회사명': 'Example Telecom',     'Industry': 'Telecommunications'},
    {'No.': 7, '한글 회사명': '샘플화재보험', '영문 회사명': 'Sample Fire Insurance', 'Industry': 'Insurance'},
]


# Marketo template column headers (row 3) and per-column markers (row 2).
TEMPLATE_HEADERS = [
    'Email Address', 'Company Name', 'First Name', 'Last Name', 'Country',
    'State', 'SFDC Campaign ID', 'Import Channel', 'Import Name', 'Import Owner',
    'Member Status', 'Channel Source', 'Channel', 'Marketing Program', 'Digital Asset',
    'Channel Team', 'Channel Team Geo', 'Initial Response Date', 'Job title', 'Job Level',
    'Department', 'Industry', 'Personal Contact Notes', 'Phone number', 'GDPR Opt-in',
    'Send GDPR follow up email', 'Requires Sales Follow-up', 'Postal Code', 'City',
    'Address', 'Website', 'Company Name (Local)', 'Salutation', 'Customer Facing Interaction',
]

TEMPLATE_MARKERS = [
    '[등록정보] 이메일가져오기', '[등록정보] 회사이름', '[등록정보] 이름', '[등록정보] 성', '고정값',
    '공란', '마케팅 제공', '드롭다운', '행사명 영문', '고정값',
    '[등록정보] 참석/불참', 'other platform', '드롭다운', '공란', '공란',
    '고정값', '고정값', '행사 날짜', '[등록정보] 한글직함', '[등록정보](LLM) 한글 직함',
    '[등록정보](LLM) 한글 부서', '[등록정보] (LLM) 한글 인더스트리', '[등록정보] 회사명/부서/직함',
    '[등록정보] 휴대폰 번호', '고정값', '공란', '[등록정보] POC 요청', '공란', '공란',
    '공란', '공란', '[등록정보] 한글로 회사 이름', '공란', '공란',
]

# Minimal Fields sheet — must be a superset of values map_industry / map_job_level / map_department return.
FIELDS_INDUSTRY = [
    'Banking and Financial Services', 'Energy and Utilities', 'Entertainment and Media',
    'Healthcare', 'Insurance', 'Life Sciences', 'Logistics', 'Manufacturing',
    'Professional Services', 'Public Sector', 'Retail', 'Technology', 'Telecommunications',
]
FIELDS_JOB_LEVEL = [
    'Executive/ C-Level', 'Vice President', 'Director', 'Manager',
    'Individual Contributor', 'Student',
]
FIELDS_DEPARTMENT = [
    'Administration and Management', 'Engineering', 'Finance', 'HR',
    'IT', 'Marketing', 'Operations', 'Sales',
]


def write_source(path):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        pd.DataFrame(SOURCE_ROWS).to_excel(writer, sheet_name='Fusion2026Data', index=False)
        pd.DataFrame(COMPANY_ROWS).to_excel(writer, sheet_name='CompanyNames', index=False)


def write_template(path, with_sample_row=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    blue = PatternFill(fill_type='solid', start_color='B4D4FF', end_color='B4D4FF')
    bold = Font(bold=True)

    # Row 1: blue spacer (Marketo's actual template puts identity / blue cells here).
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.cell(row=1, column=col_idx).fill = blue

    # Row 2: per-column source marker.
    for col_idx, marker in enumerate(TEMPLATE_MARKERS, start=1):
        ws.cell(row=2, column=col_idx, value=marker)

    # Row 3: column headers (the Marketo field names).
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = bold

    # Row 4: requirement markers (REQUIRED / Optional). Minimal version.
    for col_idx in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17):
        ws.cell(row=4, column=col_idx, value='REQUIRED')

    if with_sample_row:
        # Row 5: a sample fully-filled row so build_event_constants() can extract values.
        sample = {
            'Email Address': 'sample@example.com',
            'Company Name': 'Example Electronics',
            'First Name': 'Min',
            'Last Name': 'Kim',
            'Country': 'Korea, Republic of',
            'SFDC Campaign ID': '701Pf0000000000XXX',
            'Import Channel': 'Live Event - UiPath',
            'Import Name': 'Sample Event 2026',
            'Import Owner': 'owner@example.com',
            'Member Status': 'Attended',
            'Channel Source': 'Live Event',
            'Channel': 'Live Event - UiPath',
            'Channel Team': 'Field Marketing',
            'Channel Team Geo': 'APJ',
            'Initial Response Date': '06/16/26 09:00',
            'Job title': 'Manager',
            'Job Level': 'Manager',
            'Department': 'IT',
            'Industry': 'Technology',
            'Phone number': '01012345678',
            'GDPR Opt-in': 'Yes',
            'Company Name (Local)': '예시전자',
        }
        for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
            if header in sample:
                ws.cell(row=5, column=col_idx, value=sample[header])

    # Fields sheet for enum validation.
    fields_ws = wb.create_sheet('Fields')
    fields_ws.cell(row=1, column=1, value='Industry').font = bold
    fields_ws.cell(row=1, column=2, value='Job Level').font = bold
    fields_ws.cell(row=1, column=3, value='Department').font = bold
    for i, v in enumerate(FIELDS_INDUSTRY, start=2):
        fields_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(FIELDS_JOB_LEVEL, start=2):
        fields_ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(FIELDS_DEPARTMENT, start=2):
        fields_ws.cell(row=i, column=3, value=v)

    wb.save(path)


def main():
    source_path = os.path.join(HERE, 'sample_source.xlsx')
    template_path = os.path.join(HERE, 'sample_template.xlsx')
    prior_path = os.path.join(HERE, 'sample_prior_result.xlsx')

    write_source(source_path)
    write_template(template_path, with_sample_row=False)
    write_template(prior_path, with_sample_row=True)

    print(f'Wrote {source_path}')
    print(f'Wrote {template_path}')
    print(f'Wrote {prior_path}')


if __name__ == '__main__':
    main()
